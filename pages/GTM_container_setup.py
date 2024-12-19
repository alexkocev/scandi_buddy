import streamlit as st
import os
import json
import pandas as pd
import base64
import requests
import io
from datetime import datetime
import time
import openpyxl
import re
import json

# Load secrets
secrets_path = "./secrets.json"
with open(secrets_path, "r") as f:
    secrets = json.load(f)
CLAUDE_KEY = secrets["CLAUDE_KEY"]
url = 'https://api.anthropic.com/v1/messages'
headers = {
    'Content-Type': 'application/json',
    'x-api-key': CLAUDE_KEY,
    'anthropic-version': '2023-06-01',
    'anthropic-beta': 'pdfs-2024-09-25'
}


st.title("🦎")
st.header("GTM Container Builder")
st.markdown("💡 Upload your default container JSON and Requirements Excel file to generate a custom GTM container JSON.")

st.write("")
st.write("")

# Upload default container JSON file
col1, col2 = st.columns([3, .5])
with col1:
    default_container_file = st.file_uploader(
        "Upload the current default container JSON file",
        type=["json"],
        accept_multiple_files=False
    )

# Upload Excel requirements file
col1, col2 = st.columns([3, .5])
with col1:
    requirements_file = st.file_uploader(
        "Upload the Requirements Excel file (xlsx)",
        type=["xlsx"],
        accept_multiple_files=False,
        help="Your Excel file can have several sheets but the requirement sheet should be name 'GA4 events and parameters...'"
    )

with col2:
    st.write("")
    st.write("")
    st.write("")
    
    st.link_button(label="Help", 
                   url="https://www.loom.com/share/ec72a56dd6884b88ad5532c48b7cb4e8",
                   help="How to download the full requirement Excel file",
                   icon="🤔")
    


def fetch_matching_sheets(workbook, sheet_prefix):
    # Find all sheet names matching a prefix if needed
    matching_sheets = [sheet_name for sheet_name in workbook.sheetnames if sheet_name.startswith(sheet_prefix)]
    return matching_sheets

def process_with_retry(headers, payload, max_attempts=3, retry_delay=30):
    attempt = 0
    while attempt < max_attempts:
        try:
            response = requests.post(url, headers=headers, json=payload)
            response_json = response.json()
            if 'content' in response_json and response_json['content']:
                return response_json['content'][0]['text']
            else:
                raise ValueError(f"Unexpected response format: {response_json}")
        except Exception as e:
            attempt += 1
            if attempt < max_attempts:
                st.error(f"""Oooops we have an error, let's retry in {retry_delay} seconds... \n
                         Error on attempt {attempt}: {str(e)}""")
                time.sleep(retry_delay)
            else:
                st.error(f"""No way, max attempts reached. Come back and retry in a few minutes \n
                         Error after {max_attempts} attempts: {str(e)} """)
                return f"Error after {max_attempts} attempts: {str(e)}"


# Button to run the process
st.write("")
st.write("")
st.write("")
st.write("")

col1, col2, col3 = st.columns([1,1,1])

# Custom CSS for the button
st.markdown(
    """
    <style>
    div.stButton > button {
        background-color: #E04F4F;
        color: white;
        border: none;
        padding: 0.5em 1em;
        border-radius: 10px;
        font-size: 16px;
        font-weight: bold;
        cursor: pointer;
    }
    div.stButton > button:hover {
        background-color: #D03F3F; 
        color: white; 
    }
    </style>
    """,
    unsafe_allow_html=True
)

with col2:
    run_analysis = st.button("Build Container")

if run_analysis:
    if default_container_file is None:
        st.error("Please upload the default container JSON file.")
        st.stop()
    if requirements_file is None:
        st.error("Please upload the Requirements Excel file.")
        st.stop()

    progress_placeholder = st.empty()
    progress_placeholder.progress(0, text="Starting...")

    # Load the default container JSON
    default_container = json.load(default_container_file)

    progress_placeholder.progress(0.2, text="Reading Requirements...")

    # Load requirements from Excel
    wb = openpyxl.load_workbook(requirements_file)
    # Adjust the sheet prefix if needed, or use all sheets
    sheet_prefix = "GA4 events and parameters" # set this if needed
    matching_sheets = fetch_matching_sheets(wb, sheet_prefix)

    if not matching_sheets:
        st.warning("No matching sheets found with the prefix 'dataLayer requirements'. Using all sheets instead.")
        matching_sheets = wb.sheetnames

    requirements_list = []
    for sheet_name in matching_sheets:
        sheet_df = pd.read_excel(requirements_file, sheet_name=sheet_name)
        requirements_list.append(sheet_df)
    requirements_df = pd.concat(requirements_list, ignore_index=True)

    progress_placeholder.progress(0.4, text="Preparing prompt for Claude...")

    # Convert requirements to text
    requirements_sample = requirements_df.to_string(index=False)
    default_container_str = json.dumps(default_container, indent=2)

    best_practices_template = """
    {
        "exportFormatVersion": 2,
        "exportTime": "2024-12-16 15:51:57",
        "containerVersion": {
            "path": "accounts/ACCOUNT_ID/containers/CONTAINER_ID/versions/0",
            "accountId": "ACCOUNT_ID",
            "containerId": "CONTAINER_ID",
            "name": "GA4 EE - Base Template",
            "publicId": "GTM-XXXXX",
            "usageContext": ["WEB"],
            "tag": [
            {
                "name": "GA4 - Config",
                "type": "googtag",
                "parameter": [
                {
                    "key": "tagId",
                    "value": "{{LT - GA4 Measurement ID}}"
                }
                ],
                "firingTriggerId": ["page_view_trigger"]
            },
            {
                "name": "GA4 - EE - purchase",
                "type": "gaawe",
                "parameter": [
                {
                    "key": "eventName",
                    "value": "purchase"
                },
                {
                    "key": "eventSettingsTable",
                    "list": [
                    {"parameter": "items", "parameterValue": "{{DLV - ecommerce.items}}"},
                    {"parameter": "transaction_id", "parameterValue": "{{DLV - ecommerce.transaction_id}}"},
                    {"parameter": "value", "parameterValue": "{{DLV - ecommerce.value}}"},
                    {"parameter": "currency", "parameterValue": "{{DLV - ecommerce.currency}}"}
                    ]
                }
                ],
                "firingTriggerId": ["purchase_trigger"]
            }
            ],
            "trigger": [
            {
                "name": "Page View",
                "type": "CUSTOM_EVENT",
                "customEventFilter": [
                {
                    "type": "EQUALS",
                    "parameter": [
                    {"key": "arg0", "value": "{{_event}}"},
                    {"key": "arg1", "value": "page_view"}
                    ]
                }
                ]
            },
            {
                "name": "EE - purchase",
                "type": "CUSTOM_EVENT",
                "customEventFilter": [
                {
                    "type": "EQUALS",
                    "parameter": [
                    {"key": "arg0", "value": "{{_event}}"},
                    {"key": "arg1", "value": "purchase"}
                    ]
                }
                ]
            }
            ],
            "variable": [
            {
                "name": "DLV - ecommerce.items",
                "type": "v",
                "parameter": [
                {"key": "name", "value": "ecommerce.items"},
                {"key": "dataLayerVersion", "value": "2"}
                ]
            },
            {
                "name": "LT - GA4 Measurement ID",
                "type": "smm",
                "parameter": [
                {
                    "key": "defaultValue",
                    "value": "G-XXXXXXXX"
                }
                ]
            }
            ],
            "folder": [
            {
                "name": "GA4 - Tags",
                "containerId": "CONTAINER_ID"
            },
            {
                "name": "GA4 - EE Tags",
                "containerId": "CONTAINER_ID"
            }
            ]
        }
    }
    """
    # Prompt: Adjust as needed to instruct Claude on how to build the container
    prompt = (
        f"You are an expert in creating GTM (Google Tag Manager) container configurations for GA4 Enhanced Ecommerce events.\n"
        f"Below is the current container JSON with only default values:\n\n"
        f"### Current Container JSON:\n{default_container_str}\n\n"
        f"And here are the analytics requirements from an Excel file:\n\n"
        f"### Requirements:\n{requirements_sample}\n\n"
        f"### Below is a template container JSON that demonstrates best practices and desired structure. Use this as a reference for the final output you generate:\n\n"
        f"{best_practices_template}\n\n"
        f"Your task: Based on these requirements, modify the given default container JSON to include the necessary GA4 event tags, triggers, and variables for the events listed in the requirements.\n"
        f"Ensure that:\n"
        f"- Each required GA4 event (like page_view, view_item, add_to_cart, etc.) is implemented as a GTM GA4 event tag.\n"
        f"- Use Data Layer Variables as specified in the requirements.\n"
        f"- Maintain the same JSON structure as the provided final container template (with 'exportFormatVersion', 'containerVersion', 'tag', 'trigger', 'variable', etc.).\n"
        f"- Make sure to follow GTM best practices and the naming conventions provided in the requirements.\n"
        f"- Only return the final updated JSON container as your answer.\n"
        f"Return only the final container JSON."
    )

    progress_placeholder.progress(0.6, text="Launching the analysis...")

    payload = {
        'model': 'claude-3-5-sonnet-20241022',
        'max_tokens': 8192,
        'temperature': 0,
        'messages': [
            {
                'role': 'user',
                'content': prompt
            }
        ]
    }

    final_response = process_with_retry(headers, payload)

    if isinstance(final_response, str) and final_response.startswith("Error"):
        st.error("Failed to receive a response from Claude.")
        st.stop()

    else:
        progress_placeholder.progress(0.8, text="Parsing Claude's response...")

        # Attempt to parse the response as JSON
        try:
            container_json = json.loads(final_response)
            # If successful, provide download link as JSON
            st.success("Container JSON successfully generated!")
            container_str = json.dumps(container_json, indent=2)
            st.download_button(
                "Download GTM Container JSON",
                data=container_str,
                file_name="updated_gtm_container.json",
                mime="application/json"
            )
        except json.JSONDecodeError:
            # If not valid JSON, show the raw response
            st.warning("Claude did not return valid JSON. Showing raw response below:")
            st.text(final_response)

        progress_placeholder.progress(1.0, text="Done!")
