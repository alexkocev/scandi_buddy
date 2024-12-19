import streamlit as st
import os
import json
import pandas as pd
import base64
import requests
from datetime import datetime
import time
import openpyxl
import re
import json
from openpyxl import Workbook
from io import BytesIO
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime



# Load secrets
secrets_path = "./secrets.json"
with open(secrets_path, "r") as f:
    secrets = json.load(f)
CLAUDE_KEY = secrets["CLAUDE_KEY"]
url = 'https://api.anthropic.com/v1/messages'
headers = {
    'Content-Type': 'application/json',
    'x-api-key': CLAUDE_KEY,
    'anthropic-version': '2023-06-01',
    'anthropic-beta': 'pdfs-2024-09-25'
}


st.title("🐢")
st.header("DataLayer QA Controller")
st.markdown("💡 Upload your JSON and Excel files to generate a QA report.")

st.write("")
st.write("")

# Upload dataLayer JSON file
col1, col2 = st.columns([3, .5])
with col1:
    dataLayer_file = st.file_uploader(
        "Upload the dataLayer JSON file",
        type=["json"],
        accept_multiple_files=False
    )
with col2:
    st.write("")
    st.write("")
    st.write("")
    
    st.link_button(label="Help", 
                   url="https://www.loom.com/share/c985894788da49c085efeefe66b47e2a?sid=f96ced5c-ae7d-4bc8-a26e-148b635a2d75",
                   help="How to download the dataLayer from the Tag Assistant",
                   icon="🤔")
    
    
    
# Upload Excel requirements file
col1, col2 = st.columns([3, .5])
with col1:
    requirements_file = st.file_uploader(
        "Upload the Excel requirements file (xlsx)",
        type=["xlsx"],
        accept_multiple_files=False,
        help="Your Excel file can have several sheets but the requirement sheet should be name 'dataLayer requirements dashboard'"
    )

with col2:
    st.write("")
    st.write("")
    st.write("")
    
    st.link_button(label="Help", 
                   url="https://www.loom.com/share/ec72a56dd6884b88ad5532c48b7cb4e8",
                   help="How to download the full requirement Excel file",
                   icon="🤔")
    


def extract_json_from_response(response_text):
    match = re.search(r"\{.*\}", response_text, re.DOTALL)  # Match JSON block
    if match:
        return match.group(0)  # Return clean JSON
    else:
        return "{}"  # Return empty JSON if none found
def flatten_json(y, prefix=''):
    """Recursively flatten a nested JSON."""
    out = {}
    for key, value in y.items():
        new_key = f"{prefix}{key}" if prefix else key
        if isinstance(value, dict):
            out.update(flatten_json(value, new_key + "."))
        elif isinstance(value, list):
            for idx, item in enumerate(value):
                if isinstance(item, dict):
                    out.update(flatten_json(item, f"{new_key}[{idx}]."))
                else:
                    out[f"{new_key}[{idx}]"] = item
        else:
            out[new_key] = value
    return out

def generate_qa_excel_per_event(qa_data):
    """Generate Excel with one sheet per event."""
    wb = Workbook()
    # Remove the default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Style Definitions
    blue_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for qa_entry in qa_data:
        event_name = qa_entry['event']
        event_data = qa_entry['event_data']
        snapshot = qa_entry['snapshot']

        # Create a new sheet for the event
        sheet = wb.create_sheet(title=event_name[:30])  # Sheet names have a 30-character limit

        # Create Merged Title Row
        title = f"DataLayer QA results {datetime.now().strftime('%d.%m.%Y')}"
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
        title_cell = sheet.cell(row=1, column=1, value=title)
        title_cell.fill = blue_fill
        title_cell.font = header_font
        title_cell.alignment = center_alignment

        # Add Headers
        headers = ["Variable", "Status", "Snapshot"]
        for col_num, header in enumerate(headers, start=1):
            cell = sheet.cell(row=2, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = center_alignment

        # Write Event Data and Merge Snapshot
        row_start = 3
        rows_to_merge = len(event_data)

        # Merge Snapshot Column
        sheet.merge_cells(start_row=row_start, start_column=3, 
                          end_row=row_start + rows_to_merge - 1, end_column=3)
        snapshot_cell = sheet.cell(row=row_start, column=3, value=json.dumps(snapshot, indent=2))
        snapshot_cell.alignment = center_alignment

        # Write Variables and Statuses
        for i, entry in enumerate(event_data):
            sheet.cell(row=row_start + i, column=1, value=entry['Variable'])
            sheet.cell(row=row_start + i, column=2, value=entry['Status'])

        # Adjust Column Widths
        for col in range(1, 4):
            column_letter = get_column_letter(col)
            sheet.column_dimensions[column_letter].width = 30  # Set a reasonable default width

    return wb


def process_with_retry(headers, payload, max_attempts=3, retry_delay=30):
    attempt = 0
    while attempt < max_attempts:
        try:
            response = requests.post(url, headers=headers, json=payload)
            response_json = response.json()
            if 'content' in response_json and response_json['content']:
                return response_json['content'][0]['text']
            else:
                raise ValueError(f"Unexpected response format: {response_json}")
        except Exception as e:
            attempt += 1
            if attempt < max_attempts:
                st.error(f"""Oooops we have an error, let's retry in {retry_delay} seconds... \n
                         Error on attempt {attempt}: {str(e)}""")
                time.sleep(retry_delay)
            else:
                st.error(f"""No way, max attempts reached. Come back and retry in a few minutes \n
                         Error after {max_attempts} attempts: {str(e)} """)
                return f"Error after {max_attempts} attempts: {str(e)}"


def find_events(data, target_event):
    events = []
    # Recursive function to search for events
    def search_event(obj):
        if isinstance(obj, dict):
            # Check for the target event
            if obj.get('event') == target_event:
                events.append(obj)
            # Recursively search in dictionary values
            for key, value in obj.items():
                search_event(value)
        elif isinstance(obj, list):
            # Recursively search in list elements
            for item in obj:
                search_event(item)
    
    search_event(data)
    return events



# Run the analysis button
st.write("")
st.write("")
st.write("")
st.write("")

    
# Create a placeholder for the progress bar
progress_placeholder = st.empty()

col1, col2, col3 = st.columns([1,1,1])

# Custom CSS for the button
st.markdown(
    """
    <style>
    div.stButton > button {
        background-color: #E04F4F;
        color: white;
        border: none;
        padding: 0.5em 1em;
        border-radius: 10px;
        font-size: 16px;
        font-weight: bold;
        cursor: pointer;
    }
    div.stButton > button:hover {
        background-color: #D03F3F; 
        color: white; 
    }
    </style>
    """,
    unsafe_allow_html=True
)


with col2:
    run_analysis = st.button("Run Controller")


if run_analysis:
    if not dataLayer_file or not requirements_file:
        st.error("Please upload both the JSON and Excel files.")
        st.stop()



    # Load the JSON data
    data_layer_json = json.load(dataLayer_file)


    # Load the Excel file
    wb = openpyxl.load_workbook(requirements_file)
    sheet_names = wb.sheetnames
    # Find the index of the first sheet that starts with "dataLayer requirements"
    start_index = next((i for i, name in enumerate(sheet_names) if name.lower().startswith("datalayer requirements")), 0) + 1
    # Get all sheets after "dataLayer requirements"
    all_events = sheet_names[start_index:]

 
    # Prepare Excel Output
    output_wb = Workbook()
    del output_wb[output_wb.sheetnames[0]]  # Remove default sheet


    progress_placeholder.progress(0, text="Starting the analysis...")
    total_events = len(all_events)


    # Loop through events and analyze
    events_data = []
    qa_data = []
    
    for i, event in enumerate(all_events):
    
        progress_placeholder.progress((i + 1) / total_events, text=f"Analyzing the event '{event}'...")
        
        # Extract the relevant events from the json
        filtered_json = find_events(data_layer_json, event)

        # Check if any events were found
        if filtered_json:
            # Alex, update here to select specific event such as if login or not 🫡 
            first_events = filtered_json[:5]
            snapshot = filtered_json[0]  # Use the first event if found
            
        else:
            st.warning(f"No matching events found for '{event}'. Skipping...")
            first_events = "No Data"
            snapshot = "No Data"  # Placeholder for the snapshot

        # Extract the relevant data from the sheet.
        sheet_df = pd.read_excel(requirements_file, sheet_name=event)
        
        # Only keep the description and first table
        sheet_df = pd.concat([sheet_df.iloc[:4], sheet_df.iloc[4:].dropna(subset=['Description'])])
        for index, row in sheet_df.iterrows(): # Iterate through each row and stop when 'DataLayer' is found
            if 'DataLayer' in str(row.values):
                sheet_df = sheet_df.iloc[:index-1]
                break
            
        # Fetch list of all variables from the event
        variable_list = []
        start_collecting = False
        for index, row in sheet_df.iterrows():
            if start_collecting:
                variable_list.append(row['Description'])
            if row['Description'] == 'Variable':
                start_collecting = True
                



        # Prepare Prompt for Claude AI
        prompt = f"""
        Perform a QA on the following variables against their expected values:
        {variable_list}

        Use the following JSON snapshot for reference:
        {first_events}

        Examples of status format (only JSON format, no explanations or comments):
            "event": "Doesn't fire when clicking the \"Load More\" button on PLP (screenshot)",
            "ecommerce": "Passed QA",
            "item_list_id": "Passed QA",
            "item_name": "The value should always be picked from the English store view",
            "item_brand": "The value is in Arabic for the arabic store view. It should always be in English regardless of the picked store view",
            "price": "Passed QA",
            "battery": "The value is in Arabic for the arabic store view. It should always be in English regardless of the picked store view"

        Return only a valid JSON object with no explanations, no extra text, and no comments.
        """
        

        payload = {
            'model': 'claude-3-5-sonnet-20241022',
            'max_tokens': 8192,
            'messages': [
                {
                    'role': 'user',
                    'content': prompt
                }
            ]
        }
        
        event_data = []
        if filtered_json:
            event_response = process_with_retry(headers, payload)

            # Parse response into a DataFrame
            try:
                cleaned_response = extract_json_from_response(event_response)
                response_json = json.loads(cleaned_response)
                
            except json.JSONDecodeError:
                st.error(f"Error parsing JSON response for event {event}")
                response_json = {var: "Error" for var in variable_list}
                flattened_response = {}

            # Map variables to statuses
            for key, status in response_json.items():
                event_data.append({'Variable': key, 'Status': status})
                
        else:
            event_data = []
            for var in variable_list:
                event_data.append({"Variable": var, "Status": "No Data"})

        qa_data.append({'event': event, 'event_data': event_data, 'snapshot': snapshot})


        

    
    
    



    st.success("QA Process Completed!")



    excel_wb = generate_qa_excel_per_event(qa_data)
    output_stream = BytesIO()
    excel_wb.save(output_stream)
    output_stream.seek(0)

    st.download_button(
        label="Download QA Report",
        data=output_stream,
        file_name="datalayer_qa_output.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )  
    