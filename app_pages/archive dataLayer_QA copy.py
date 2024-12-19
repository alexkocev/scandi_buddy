import streamlit as st
import os
import json
import pandas as pd
import base64
import requests
import io
from datetime import datetime
import time
import openpyxl
import re
import json
from dotenv import load_dotenv

# Load environment variables from .env file (optional for local dev)
load_dotenv()

# Load secrets
CLAUDE_KEY = os.getenv("CLAUDE_KEY_SW")
url = 'https://api.anthropic.com/v1/messages'
headers = {
    'Content-Type': 'application/json',
    'x-api-key': CLAUDE_KEY,
    'anthropic-version': '2023-06-01',
    'anthropic-beta': 'pdfs-2024-09-25'
}


st.title("🐢")
st.header("DataLayer QA Controler")
st.markdown("💡 Upload your JSON and Excel files to generate a QA report.")

st.write("")
st.write("")

# Upload dataLayer JSON file
col1, col2 = st.columns([3, .5])
with col1:
    dataLayer_file = st.file_uploader(
        "Upload the dataLayer JSON file",
        type=["json"],
        accept_multiple_files=False
    )

# Upload Excel requirements file
col1, col2 = st.columns([3, .5])
with col1:
    requirements_file = st.file_uploader(
        "Upload the Excel requirements file (xlsx)",
        type=["xlsx"],
        accept_multiple_files=False,
        help="Your Excel file can have several sheets but the requirement sheet should be name 'dataLayer requirements ...'"
    )

# Custom prompt parameters (optional)
with st.expander("⚙️ Additional Parameters", expanded=False):
    custom_request = st.text_area(
        "Custom request (optional)",
        value="",
        placeholder="Add any custom instructions here."
    )






def extract_unique_events_from_json(data):
    # Extract unique events from the provided JSON data
    unique_events = []
    seen_events = set()

    for container in data.get('data', {}).get('containers', []):
        for message in container.get('messages', []):
            event_data = {
                "eventName": message.get("eventName"),
                "tagInfo": message.get("tagInfo"),
                "consentData": message.get("consentData"),
                "ruleInfo": message.get("ruleInfo"),
                "message": message.get("message")
            }
            event_data_str = json.dumps(event_data, sort_keys=True)
            if event_data_str not in seen_events:
                seen_events.add(event_data_str)
                unique_events.append(event_data)
    return unique_events


def fetch_matching_sheets(workbook, sheet_prefix):
    # Find all sheet names matching the prefix
    matching_sheets = [sheet_name for sheet_name in workbook.sheetnames if sheet_name.startswith(sheet_prefix)]
    return matching_sheets




# Run the analysis button
st.write("")
st.write("")
st.write("")
st.write("")

def process_with_retry(headers, payload, max_attempts=3, retry_delay=30):
    attempt = 0
    while attempt < max_attempts:
        try:
            response = requests.post(url, headers=headers, json=payload)
            response_json = response.json()
            if 'content' in response_json and response_json['content']:
                return response_json['content'][0]['text']
            else:
                raise ValueError(f"Unexpected response format: {response_json}")
        except Exception as e:
            attempt += 1
            if attempt < max_attempts:
                st.error(f"""Oooops we have an error, let's retry in {retry_delay} seconds... \n
                         Error on attempt {attempt}: {str(e)}""")
                time.sleep(retry_delay)
            else:
                st.error(f"""No way, max attempts reached. Come back and retry in a few minutes \n
                         Error after {max_attempts} attempts: {str(e)} """)
                return f"Error after {max_attempts} attempts: {str(e)}"

# Create a placeholder for the progress bar
progress_placeholder = st.empty()

col1, col2, col3 = st.columns([1,1,1])

# Custom CSS for the button
st.markdown(
    """
    <style>
    div.stButton > button {
        background-color: #E04F4F;
        color: white;
        border: none;
        padding: 0.5em 1em;
        border-radius: 10px;
        font-size: 16px;
        font-weight: bold;
        cursor: pointer;
    }
    div.stButton > button:hover {
        background-color: #D03F3F; 
        color: white; 
    }
    </style>
    """,
    unsafe_allow_html=True
)


with col2:
    run_analysis = st.button("Run Analysis")


if run_analysis:
    if dataLayer_file is None:
        st.error("Please upload the dataLayer JSON file.")
        st.stop()
    if requirements_file is None:
        st.error("Please upload the Excel requirements file.")
        st.stop()

    progress_placeholder.progress(0, text="Starting the analysis...")

    # Load the JSON data
    data_layer_json = json.load(dataLayer_file)
    extracted_events = extract_unique_events_from_json(data_layer_json)

    progress_placeholder.progress(0.2, text="Converting Json and Excel files...")

    # Load requirements from Excel
    wb = openpyxl.load_workbook(requirements_file)
    # Identify relevant sheets (if needed)
    # For this example, we assume the main requirements are on a sheet that starts with "dataLayer requirements"
    sheet_prefix = "dataLayer requirements"
    matching_sheets = fetch_matching_sheets(wb, sheet_prefix)

    if not matching_sheets:
        st.warning("No matching sheets found with the prefix 'dataLayer requirements'. Using all sheets instead.")
        matching_sheets = wb.sheetnames

    # Combine all matching sheets into a single DataFrame
    requirements_list = []
    for sheet_name in matching_sheets:
        sheet_df = pd.read_excel(requirements_file, sheet_name=sheet_name)
        requirements_list.append(sheet_df)
    requirements_df = pd.concat(requirements_list, ignore_index=True)
                
    progress_placeholder.progress(0.5, text="Launching the analysis...")


    # Prepare prompt for Claude
    example_output = (
        "dataLayer events\tStatus\tDescription\tExpected places for triggering\tNOTES\n"
        "view_item_list\tImprovement opportunity\tView of item impressions in a list.\t"
        "Log this event when the user has been presented with a list of items...\t"
        "\"item_brand\" is missing...\n"
        "view_item\tCollecting data correctly\tView item details.\tevent fires when accessing productDetails page\t\n"
        "add_to_cart\tCollecting data correctly\tAdd item(s) to cart.\tevent fires when adding a product...\t\n"
        "remove_from_cart\tCollecting data correctly\tRemove item(s) from the cart.\tevent fires when removing...\t\n"
    )

    extracted_events_sample = json.dumps(extracted_events, indent=2)
    requirements_sample = requirements_df.to_string(index=False)

    prompt = (
        f"You are tasked with performing a DataLayer QA for a website's analytics setup. "
        f"The goal is to map events extracted from GTM (Google Tag Manager) against the provided analytics setup requirements.\n\n"
        f"### Task:\n"
        f"1. List **all events** from the requirements file, even if they are not found in the extracted GTM events.\n"
        f"2. If an event is **not found** in the extracted events, mark the 'Status' column as 'More Work Needed' and write 'Event not found' in the 'NOTES' column.\n"
        f"3. If an event is present in the extracted events, analyze it and choose the correct status: \n"
        f"   - 'Collecting data correctly': The event is correctly implemented.\n"
        f"   - 'Improvement opportunity': The event fires, but improvements are needed (explain in NOTES).\n"
        f"   - 'QA on Production': The event is partially implemented and requires production testing.\n"
        f"   - 'Out of scope': Non-relevant events.\n"
        f"   - 'More Work Needed': The event is missing or incomplete.\n\n"
        f"### Example Output:\n{example_output}\n\n"
        f"### Requirements:\n{requirements_sample}\n\n"
        f"### Extracted Events:\n{extracted_events_sample}\n\n"
        f"{f'Additional instructions: {custom_request}' if custom_request else ''}\n\n"
        f"### Expected Output:\n"
        f"Provide the output in a table format with the following columns:\n"
        f"1. `dataLayer events`\n"
        f"2. `Status`\n"
        f"3. `Description`\n"
        f"4. `Expected places for triggering`\n"
        f"5. `NOTES`\n\n"
        f"Ensure all events from the requirements file are included, with appropriate statuses and notes."
    )

    payload = {
        'model': 'claude-3-5-sonnet-20241022',
        'max_tokens': 8192,
        'messages': [
            {
                'role': 'user',
                'content': prompt
            }
        ]
    }
    
    
    

    final_response = process_with_retry(headers, payload)

    if isinstance(final_response, str) and final_response.startswith("Error"):
        st.error("Failed to receive a response from Claude.")
        st.stop()
    else:
        # Claude's response is expected to be a string with table format
        # Extract 'content' from response
        # response should be a list of messages. We'll parse the first one.
        
        progress_placeholder.progress(1.0, text="Generating PDF report...")

        qa_results = final_response
        
        
        # Convert response to a DataFrame
        rows = [row.split('|') for row in qa_results.splitlines() if '|' in row]
        headers = rows[0]  # Extract headers from the first row
        data = rows[1:]    # Extract the data rows
        
        qa_df = pd.DataFrame(data, columns=[col.strip() for col in headers])

        # Create an output file in memory
        output_buffer = io.BytesIO()
        with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
            qa_df.to_excel(writer, index=False)

        st.success("QA Analysis Completed!")
        st.download_button(
            "Download QA Results Excel",
            data=output_buffer.getvalue(),
            file_name="datalayer_qa_output.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    
    
