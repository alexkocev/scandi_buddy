import streamlit as st
import os
import json
import pandas as pd
import base64
import requests
from datetime import datetime
import time
import openpyxl
import re
from openpyxl import Workbook
from io import BytesIO
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
import subprocess
from pathlib import Path
from dotenv import load_dotenv
from google.cloud import storage
from google.oauth2 import service_account
    
# Load environment variables from .env file (optional for local dev)
load_dotenv()


# Load secrets
CLAUDE_KEY = os.getenv("CLAUDE_KEY_SW")
url = 'https://api.anthropic.com/v1/messages'
headers = {
    'Content-Type': 'application/json',
    'x-api-key': CLAUDE_KEY,
    'anthropic-version': '2023-06-01',
    'anthropic-beta': 'pdfs-2024-09-25'
}


st.title("🐢")
st.header("DataLayer QA Controller")
st.markdown("💡 Simulate your actions or upload JSON and Excel files to generate a QA report.")

st.write("")
st.write("")








# -- Storage --


# Function to upload file to GCS with a timestamped name
def upload_to_gcs(file, bucket_name, destination_blob_name, credentials_info):
    # Create credentials object
    credentials = service_account.Credentials.from_service_account_info(credentials_info)
    # Initialize the storage client with credentials
    client = storage.Client(credentials=credentials)
    bucket = client.bucket(bucket_name)
    blob = bucket.blob(destination_blob_name)
    blob.upload_from_file(file)
    return f"gs://{bucket_name}/{destination_blob_name}"

# Function to download the file temporarily
def download_file_from_gcs(bucket_name, blob_name, local_file_path, credentials_info):
    # Create credentials object
    credentials = service_account.Credentials.from_service_account_info(credentials_info)
    # Initialize the storage client with credentials
    client = storage.Client(credentials=credentials)
    bucket = client.bucket(bucket_name)
    blob = bucket.blob(blob_name)
    blob.download_to_filename(local_file_path)
    return local_file_path

# Function to read file content from GCS
def read_file_from_gcs(bucket_name, blob_name, credentials_info):
    credentials = service_account.Credentials.from_service_account_info(credentials_info)
    client = storage.Client(credentials=credentials)
    bucket = client.bucket(bucket_name)
    blob = bucket.blob(blob_name)

    # Read the file as text (assuming JSON content)
    file_content = blob.download_as_text(encoding="utf-8")
    return file_content



    
    
    



    

# -- TRACKING USER ACTION (optional) --
# Placeholder for the JSON data
if 'interaction_data' not in st.session_state:
    st.session_state.interaction_data = []



# Generate a timestamp for unique file naming
def generate_timestamp():
    return datetime.now().strftime("%Y%m%d_%H%M%S")

# Function to run the Playwright script
def run_script(url, timestamp):
    script_path = "./scripts/dataLayer_QA/interaction_tracker.py"  # Path to your Playwright script
    chrome_process = subprocess.Popen(
        ["python", script_path, url, timestamp],
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        universal_newlines=True
    )
    return chrome_process

# Function to fetch interaction data
def fetch_interaction_data(timestamp):
    json_file = Path(f"data/dataLayer_QA/temp_json/interaction_data_{timestamp}.json")
    if json_file.exists():
        with json_file.open("r", encoding="utf-8") as f:
            st.session_state.interaction_data = json.load(f)
            return st.session_state.interaction_data
    else:
        st.warning("No interaction data file found!")
        return []

col1, col2 = st.columns([3, .5])
with col1:
    st.markdown("<small>Enter the URL and start the simulation: (optional) 🚨 Need to run locally</small>", unsafe_allow_html=True)
    interaction_tracker_container = st.container(border=True)
    with interaction_tracker_container:
        container_col1, container_col2 = st.columns([3, .8])

        with container_col1:
            # Input for the URL
            simulation_url = st.text_input("Enter the URL and start the simulation: (optional) 🚨 Need to run locally", 
                                placeholder="https://example.com", 
                                help="Provide the website URL you want to simulate interactions on. Leave blank if not needed.",
                                label_visibility="collapsed")

        # Generate a unique timestamp for each session
        timestamp = generate_timestamp()
        
        # Create a button to trigger the analysis
        with container_col2:            
            run_tracking = st.button("Run Tracking", help="Start the interaction simulation")

        if run_tracking:
            if simulation_url and simulation_url.startswith("http"):
                chrome_process = run_script(simulation_url, timestamp)
                st.info("Chrome launched. Interact with the browser. Close the Chrome window when done.")
                
                # Monitor the Chrome process
                while chrome_process.poll() is None:
                    time.sleep(1)  # Wait for the process to complete
                
                # Once Chrome window is closed
                interaction_data = fetch_interaction_data(timestamp)
                if interaction_data:
                    st.success("Simulation done and data saved successfully!")
                    with st.expander("See the interaction data"):
                        st.code(json.dumps(interaction_data, indent=4))
                else:
                    st.warning("No data recorded during the session.")
            else:
                st.error("Please provide a valid URL.")

















# Upload dataLayer JSON file
col1, col2 = st.columns([3, .5])
with col1:
    dataLayer_file = st.file_uploader(
        "Upload the dataLayer JSON file 🚨 Need to run locally if more than 32Mb",
        type=["json"],
        accept_multiple_files=False,
        help="Upload the exported JSON file from the Tag Assistant containing website interaction data for QA."
    )


            
with col2:
    st.write("")
    st.write("")
    st.write("")
    
    st.link_button(label="Help", 
                   url="https://www.loom.com/share/c985894788da49c085efeefe66b47e2a?sid=f96ced5c-ae7d-4bc8-a26e-148b635a2d75",
                   help="How to download the dataLayer from the Tag Assistant",
                   icon="🤔")
    
    
    
# Upload Excel requirements file
col1, col2 = st.columns([3, .5])
with col1:
    requirements_file = st.file_uploader(
        "Upload the Excel requirements file (xlsx)",
        type=["xlsx"],
        accept_multiple_files=False,
        help="Your Excel file can have several sheets but the requirement sheet should be name 'dataLayer requirements dashboard'"
    )

with col2:
    st.write("")
    st.write("")
    st.write("")
    
    st.link_button(label="Help", 
                   url="https://www.loom.com/share/ec72a56dd6884b88ad5532c48b7cb4e8",
                   help="How to download the full requirement Excel file",
                   icon="🤔")
    


def extract_json_from_response(response_text):
    match = re.search(r"\{.*\}", response_text, re.DOTALL)  # Match JSON block
    if match:
        return match.group(0)  # Return clean JSON
    else:
        return "{}"  # Return empty JSON if none found
def flatten_json(y, prefix=''):
    """Recursively flatten a nested JSON."""
    out = {}
    for key, value in y.items():
        new_key = f"{prefix}{key}" if prefix else key
        if isinstance(value, dict):
            out.update(flatten_json(value, new_key + "."))
        elif isinstance(value, list):
            for idx, item in enumerate(value):
                if isinstance(item, dict):
                    out.update(flatten_json(item, f"{new_key}[{idx}]."))
                else:
                    out[f"{new_key}[{idx}]"] = item
        else:
            out[new_key] = value
    return out

def generate_qa_excel_per_event(qa_data):
    """Generate Excel with one sheet per event."""
    wb = Workbook()
    # Remove the default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Style Definitions
    blue_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for qa_entry in qa_data:
        event_name = qa_entry['event']
        event_data = qa_entry['event_data']
        snapshot = qa_entry['snapshot']

        # Create a new sheet for the event
        sheet = wb.create_sheet(title=event_name[:30])  # Sheet names have a 30-character limit

        # Create Merged Title Row
        title = f"DataLayer QA results {datetime.now().strftime('%d.%m.%Y')}"
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
        title_cell = sheet.cell(row=1, column=1, value=title)
        title_cell.fill = blue_fill
        title_cell.font = header_font
        title_cell.alignment = center_alignment

        # Add Headers
        headers = ["Variable", "Status", "Snapshot"]
        for col_num, header in enumerate(headers, start=1):
            cell = sheet.cell(row=2, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = center_alignment

        # Write Event Data and Merge Snapshot
        row_start = 3
        rows_to_merge = len(event_data)

        # Merge Snapshot Column
        sheet.merge_cells(start_row=row_start, start_column=3, 
                          end_row=row_start + rows_to_merge - 1, end_column=3)
        snapshot_cell = sheet.cell(row=row_start, column=3, value=json.dumps(snapshot, indent=2))
        snapshot_cell.alignment = center_alignment

        # Write Variables and Statuses
        for i, entry in enumerate(event_data):
            sheet.cell(row=row_start + i, column=1, value=entry['Variable'])
            sheet.cell(row=row_start + i, column=2, value=entry['Status'])

        # Adjust Column Widths
        for col in range(1, 4):
            column_letter = get_column_letter(col)
            sheet.column_dimensions[column_letter].width = 30  # Set a reasonable default width

    return wb


def process_with_retry(headers, payload, max_attempts=3, retry_delay=30):
    attempt = 0
    while attempt < max_attempts:
        try:
            response = requests.post(url, headers=headers, json=payload)
            response_json = response.json()
            if 'content' in response_json and response_json['content']:
                return response_json['content'][0]['text']
            else:
                raise ValueError(f"Unexpected response format: {response_json}")
        except Exception as e:
            attempt += 1
            if attempt < max_attempts:
                st.error(f"""Oooops we have an error, let's retry in {retry_delay} seconds... \n
                         Error on attempt {attempt}: {str(e)}""")
                time.sleep(retry_delay)
            else:
                st.error(f"""No way, max attempts reached. Come back and retry in a few minutes \n
                         Error after {max_attempts} attempts: {str(e)} """)
                return f"Error after {max_attempts} attempts: {str(e)}"


def find_events(data, target_event):
    events = []
    # Recursive function to search for events
    def search_event(obj):
        if isinstance(obj, dict):
            # Check for the target event
            if obj.get('event') == target_event:
                events.append(obj)
            # Recursively search in dictionary values
            for key, value in obj.items():
                search_event(value)
        elif isinstance(obj, list):
            # Recursively search in list elements
            for item in obj:
                search_event(item)
    
    search_event(data)
    return events



# Run the analysis button
st.write("")
st.write("")
st.write("")
st.write("")

    
# Create a placeholder for the progress bar
progress_placeholder = st.empty()

col1, col2, col3 = st.columns([1,1,1])

# Custom CSS for the button
st.markdown(
    """
    <style>
    div.stButton > button {
        background-color: #E04F4F;
        color: white;
        border: none;
        padding: 0.5em 1em;
        border-radius: 10px;
        font-size: 16px;
        font-weight: bold;
        cursor: pointer;
    }
    div.stButton > button:hover {
        background-color: #D03F3F; 
        color: white; 
    }
    </style>
    """,
    unsafe_allow_html=True
)


with col2:
    run_analysis = st.button("Run Controller")


if run_analysis:
    if not dataLayer_file or not requirements_file:
        st.error("Please upload both the JSON and Excel files.")
        st.stop()



    # Load the JSON data
    data_layer_json = json.load(dataLayer_file)


    # Load the Excel file
    wb = openpyxl.load_workbook(requirements_file)
    sheet_names = wb.sheetnames
    # Find the index of the first sheet that starts with "dataLayer requirements"
    start_index = next((i for i, name in enumerate(sheet_names) if name.lower().startswith("datalayer requirements")), 0) + 1
    # Get all sheets after "dataLayer requirements"
    all_events = sheet_names[start_index:]

 
    # Prepare Excel Output
    output_wb = Workbook()
    del output_wb[output_wb.sheetnames[0]]  # Remove default sheet


    progress_placeholder.progress(0, text="Starting the analysis...")
    total_events = len(all_events)


    # Loop through events and analyze
    events_data = []
    qa_data = []
    
    for i, event in enumerate(all_events):
    
        progress_placeholder.progress((i + 1) / total_events, text=f"Analyzing the event '{event}'...")
        
        # Extract the relevant events from the json
        filtered_json = find_events(data_layer_json, event)
        
        # Check if any events were found
        if filtered_json:
            # Alex, update here to select specific event such as if login or not 🫡 
            first_events = filtered_json[:10]
            snapshot = filtered_json[0]  # Use the first event if found
            
        else:
            st.warning(f"No matching events found for '{event}'. Skipping...")
            first_events = "No Data"
            snapshot = "No Data"  # Placeholder for the snapshot

        # Extract the relevant data from the sheet.
        sheet_df = pd.read_excel(requirements_file, sheet_name=event)
        
        # Only keep the description and first table
        sheet_df = pd.concat([sheet_df.iloc[:4], sheet_df.iloc[4:].dropna(subset=['Description'])])
        for index, row in sheet_df.iterrows(): # Iterate through each row and stop when 'DataLayer' is found
            if 'DataLayer' in str(row.values):
                sheet_df = sheet_df.iloc[:index-1]
                break
            
        # Fetch list of all variables from the event
        variable_list = []
        start_collecting = False
        for index, row in sheet_df.iterrows():
            if start_collecting:
                variable_list.append(row['Description'])
            if row['Description'] == 'Variable':
                start_collecting = True
                

        interaction_data_section = ""
        if st.session_state.interaction_data:
            interaction_data_section = f"""
            Additionally, consider the following user interaction data collected during the simulation:
            {json.dumps(st.session_state.interaction_data, indent=4)}
            """
    
        # Prepare Prompt for Claude AI
        prompt = f"""
        Perform a QA on the following variables against their expected values:
        {variable_list}

        Use the following JSON snapshot for reference:
        {first_events}
        
        {interaction_data_section}

        ----- 
        ## Examples of status format (only JSON format, no explanations or comments):

        Examples for the event view_item_list:
        "event": "Passed QA",
        "item_brand": "Missing for the products on Home Page. Otherwise all good",
        "item_type": "Not found",
        "item_stock_status": "Value not found even though it was enabled in the admin panel",
        "item_sale_product": "Value not found even though it was enabled in the admin panel",
        "item_list_id": "On AR store view, the parameter value is in AR, while it should be in EN",
        "battery": "The parameter value is 'N/A'; ensure values are fetched from EN store view"

        Examples for the event login
        "event": "When the user signs up, they log in automatically. We should also fire a login event",
        "method": "Isn't dynamic. Should pass a value related to the used login method (Google/Facebook/Email). It always passes Email"

        Examples for the event cartStatus
        "price": "Price value doesn't change in case item is on discount and still shows original price",
        "brand": "Brand value is 'N/A' in 'cartStatus' event"

        Examples for the event purchase
        "discount_amount": "The value is negative",
        "coupon": "The parameter is not populated",
        "customer_name": "Missing from the dataLayer when a user is not logged in",
        "brand": "Missing from the dataLayer",
        "currency": "Present in the ecommerce object, but not in the ITEMS array. Should be added in both places",
        "event": [
        "Fires twice; disable the old module",
        "Fires for uWallet even though the transaction is not completed"
        ],
        "item_brand": "The value is in Arabic for the Arabic store view. It should always be in English regardless of the picked storeview"

        Examples for the event impressions
        "variant": "Variant value is empty in impression event",
        "list": "List parameter is not present"

        Examples for the event productClick
        "category": "Category value is empty in productClick event",
        "brand": "Its value is 'N/A' instead of the brand's name"

        Examples for the event productDetail
        "brand": "Brand value is empty in productDetail event",
        "dimension2": "Simple product Name of added product (only for Configurable products, when there is information about Simple product available, in other cases pass "N/A"). Please reffer to Sheet "ID and Name" event "For some products in cart and minicart when changing product quantity can be observed error and quantity is not possible to change https://share.getcloudapp.com/GGuWv7eZ . \n But from the PDP page product can be normally added to cart. \n The issue product “L-Citrullin Kapseln” can be found under menu GESUNDHEIT - Herz & Kreislauf"


        Examples for the event addToCart
        "quantity": "Should represent how many items have been added individually, not total similar products in the cart",
        "event": "Couldn't add product to cart from PLP due to UI bug",
        "variant": "Value is 'N/A' instead of the correct variant",
        "currency": "Present in the ecommerce object, but not in the ITEMS array. Should be added in both places"

        Examples for the event checkout
        "step 1": "Should fire on View Cart page instead of when starting checkout. The event sequence should also be changed",
        "step 2": [
        "Should be labeled as step 3 (payment page)",
        "Brand value is 'N/A' in 'checkout' event"
        ]

        Examples for the event general
        "store_view": "Passed QA",
        "pageType": [
        "Passed QA",
        "Should represent 'success' when a user places a purchase, not 'checkout'"
        ],
        "user_id": [
        "Passed QA",
        "Remains null when a user logs into the homepage",
        "Null is being passed as a string, not as null"
        ],
        "event": [
        "Firing slightly too late; it should fire first when the page loads",
        "Multiple redundant events firing before DOM load"
        ],
        "event": "Fires regardless of registration success",
        "customerEmail": "Sent as 'N/A'",
        "message": "Sometimes empty, making the functionality unclear",
        "filterCategory": "Does not show up if STORE LOCATION filter is applied"
        
        ------

        Return only a valid JSON object with no explanations, no extra text, and no comments.
        """
        

        payload = {
            'model': 'claude-3-5-sonnet-20241022',
            'max_tokens': 8192,
            'messages': [
                {
                    'role': 'user',
                    'content': prompt
                }
            ]
        }
        
        event_data = []
        if filtered_json:
            event_response = process_with_retry(headers, payload)

            # Parse response into a DataFrame
            try:
                cleaned_response = extract_json_from_response(event_response)
                response_json = json.loads(cleaned_response)
                
            except json.JSONDecodeError:
                st.error(f"Error parsing JSON response for event {event}")
                response_json = {var: "Error" for var in variable_list}
                flattened_response = {}

            # Map variables to statuses
            for key, status in response_json.items():
                event_data.append({'Variable': key, 'Status': status})
                
        else:
            event_data = []
            for var in variable_list:
                event_data.append({"Variable": var, "Status": "No Data"})

        qa_data.append({'event': event, 'event_data': event_data, 'snapshot': snapshot})


        

    
    
    



    st.success("QA Process Completed!")



    excel_wb = generate_qa_excel_per_event(qa_data)
    output_stream = BytesIO()
    excel_wb.save(output_stream)
    output_stream.seek(0)

    st.download_button(
        label="Download QA Report",
        data=output_stream,
        file_name="datalayer_qa_output.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )  
    